from fastapi import FastAPI
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import google.generativeai as genai

# ==================================
# GEMINI
# ==================================

API_KEY = "AQ.Ab8RN6InMsgbz9z2nab4JSnNDgvYWpwd8qJ5wcsXx0yXdQ56cg"

genai.configure(api_key=API_KEY)

model = genai.GenerativeModel("gemini-2.5-flash")

# ==================================
# FASTAPI
# ==================================

app = FastAPI()

# ==================================
# EXCEL
# ==================================

EXCEL_FILE = "appointments.xlsx"

if not os.path.exists(EXCEL_FILE):

    workbook = Workbook()

    sheet = workbook.active

    sheet.append([
        "Appointment ID",
        "Created At",
        "Name",
        "Phone",
        "Date",
        "Time",
        "Status"
    ])

    workbook.save(EXCEL_FILE)

# ==================================
# SESSION MEMORY
# ==================================

sessions = {}

# ==================================
# MODELS
# ==================================

class ChatRequest(BaseModel):
    session_id: str
    message: str

class Appointment(BaseModel):
    name: str
    phone: str
    date: str
    time: str

# ==================================
# HELPERS
# ==================================

def generate_id():

    return (
        "GYR-"
        + datetime.now().strftime("%Y%m%d%H%M%S")
    )

def save_appointment(
    appointment_id,
    name,
    phone,
    date,
    time
):

    workbook = load_workbook(EXCEL_FILE)

    sheet = workbook.active

    sheet.append([
        appointment_id,
        datetime.now().strftime(
            "%d-%m-%Y %I:%M %p"
        ),
        name,
        phone,
        date,
        time,
        "Scheduled"
    ])

    workbook.save(EXCEL_FILE)

def interpret_datetime(text):

    prompt = f"""
Today's date is {datetime.now()}.

Convert this appointment request into:

DATE=DD-MM-YYYY
TIME=HH:MM AM/PM

Request:
{text}

Examples:

tomorrow evening at 6

DATE=08-06-2026
TIME=06:00 PM

next monday at 4

DATE=15-06-2026
TIME=04:00 PM

first week of july around 6

DATE=03-07-2026
TIME=06:00 PM

Return ONLY:

DATE=...
TIME=...
"""

    response = model.generate_content(prompt)

    result = response.text.strip()

    lines = result.split("\n")

    date = (
        lines[0]
        .replace("DATE=", "")
        .strip()
    )

    time = (
        lines[1]
        .replace("TIME=", "")
        .strip()
    )

    return date, time

# ==================================
# ROUTES
# ==================================

@app.get("/")
def home():

    return {
        "message":
        "Susan AI Receptionist Running"
    }

@app.post("/appointment")
def create_appointment(
    appointment: Appointment
):

    appointment_id = generate_id()

    save_appointment(
        appointment_id,
        appointment.name,
        appointment.phone,
        appointment.date,
        appointment.time
    )

    return {

        "status": "success",

        "appointment_id":
        appointment_id
    }

# ==================================
# SUSAN CHAT
# ==================================

@app.post("/chat")
def chat(request: ChatRequest):

    session_id = request.session_id

    message = request.message

    if session_id not in sessions:

        sessions[session_id] = {
            "step": "start"
        }

    session = sessions[session_id]

    # ===============================
    # START
    # ===============================

    if session["step"] == "start":

        if (
            "appointment" in message.lower()
            or "book" in message.lower()
        ):

            session["step"] = "name"

            return {
                "reply":
                "Certainly. May I have your full name?"
            }

        prompt = f"""
You are Susan,
receptionist of Gyrobot.

Customer:
{message}
"""

        response = model.generate_content(
            prompt
        )

        return {
            "reply":
            response.text
        }

    # ===============================
    # NAME
    # ===============================

    elif session["step"] == "name":

        session["name"] = message

        session["step"] = "phone"

        return {
            "reply":
            "Thank you. Could I have your phone number?"
        }

    # ===============================
    # PHONE
    # ===============================

    elif session["step"] == "phone":

        session["phone"] = message

        session["step"] = "datetime"

        return {
            "reply":
            "What date and time would you prefer?"
        }

    # ===============================
    # DATE TIME
    # ===============================

    elif session["step"] == "datetime":

        date, time = interpret_datetime(
            message
        )

        session["date"] = date
        session["time"] = time

        session["step"] = "confirm"

        return {
            "reply":
            f"""
Please confirm:

Name: {session['name']}
Phone: {session['phone']}
Date: {date}
Time: {time}

Reply YES to confirm.
"""
        }

    # ===============================
    # CONFIRM
    # ===============================

    elif session["step"] == "confirm":

        if message.lower() == "yes":

            appointment_id = generate_id()

            save_appointment(
                appointment_id,
                session["name"],
                session["phone"],
                session["date"],
                session["time"]
            )

            del sessions[session_id]

            return {
                "reply":
                f"""
Perfect.

Your appointment has been booked.

Appointment ID:
{appointment_id}
"""
            }

        else:

            del sessions[session_id]

            return {
                "reply":
                "Appointment cancelled."
            }

