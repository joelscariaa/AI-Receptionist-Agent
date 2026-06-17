import google.generativeai as genai
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# =====================================
# CONFIGURATION
# =====================================

API_KEY = "AQ.Ab8RN6InMsgbz9z2nab4JSnNDgvYWpwd8qJ5wcsXx0yXdQ56cg"

genai.configure(api_key=API_KEY)

model = genai.GenerativeModel("gemini-2.5-flash")

EXCEL_FILE = "appointments.xlsx"

# =====================================
# CREATE EXCEL FILE
# =====================================

if not os.path.exists(EXCEL_FILE):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Appointments"

    sheet.append([
        "Appointment ID",
        "Booking Timestamp",
        "Customer Name",
        "Phone Number",
        "Appointment Date",
        "Appointment Time"
    ])

    workbook.save(EXCEL_FILE)

# =====================================
# GENERATE APPOINTMENT ID
# =====================================

def generate_appointment_id():

    return "GYR-" + datetime.now().strftime("%Y%m%d%H%M%S")

# =====================================
# SAVE APPOINTMENT
# =====================================

def save_appointment(
    appointment_id,
    name,
    phone,
    date,
    time
):

    workbook = load_workbook(EXCEL_FILE)

    sheet = workbook.active

    sheet.append([
        appointment_id,
        datetime.now().strftime("%d-%m-%Y %I:%M %p"),
        name,
        phone,
        date,
        time
    ])

    workbook.save(EXCEL_FILE)

# =====================================
# INTERPRET DATE & TIME
# =====================================

def interpret_appointment_datetime(user_input):

    prompt = f"""
Today's date is {datetime.now().strftime('%d-%m-%Y')}.

Interpret the appointment request.

Examples:

tomorrow at 6 in the evening
DATE=08-06-2026
TIME=06:00 PM

next monday evening
DATE=15-06-2026
TIME=06:00 PM

first week of july around 4
DATE=03-07-2026
TIME=04:00 PM

in 2 hours
DATE=today
TIME=current time + 2 hours

User Request:
{user_input}

Return ONLY:

DATE=DD-MM-YYYY
TIME=HH:MM AM/PM
"""

    try:

        response = model.generate_content(prompt)

        result = response.text.strip()

        lines = result.split("\n")

        final_date = lines[0].replace(
            "DATE=",
            ""
        ).strip()

        final_time = lines[1].replace(
            "TIME=",
            ""
        ).strip()

        return final_date, final_time

    except Exception:

        return None, None

# =====================================
# START RECEPTIONIST
# =====================================

print("\nSusan: Hello! Welcome to Gyrobot.")
print("Susan: How may I assist you today?")
print("Type 'exit' to quit.\n")

while True:

    user_message = input("You: ")

    if user_message.lower() == "exit":

        print(
            "\nSusan: Thank you for contacting Gyrobot. Goodbye!"
        )

        break

    # =================================
    # APPOINTMENT BOOKING
    # =================================

    if (
        "appointment" in user_message.lower()
        or "book" in user_message.lower()
    ):

        print(
            "\nSusan: Certainly. May I have your full name?"
        )

        name = input("You: ")

        print(
            "\nSusan: Thank you. Could I have your phone number?"
        )

        phone = input("You: ")

        print(
            "\nSusan: What date and time would you prefer?"
        )

        print(
            "Susan: Example: next monday evening at 6"
        )

        appointment_request = input("You: ")

        final_date, final_time = (
            interpret_appointment_datetime(
                appointment_request
            )
        )

        if final_date is None:

            print(
                "\nSusan: Sorry, I couldn't understand the appointment time."
            )

            print(
                "Susan: Please try again using a clearer date and time."
            )

            continue

        print("\nSusan: Please confirm the details.\n")

        print(f"Name : {name}")
        print(f"Phone: {phone}")
        print(f"Date : {final_date}")
        print(f"Time : {final_time}")

        confirmation = input(
            "\nSusan: Confirm appointment? (yes/no): "
        )

        if confirmation.lower() == "yes":

            appointment_id = (
                generate_appointment_id()
            )

            save_appointment(
                appointment_id,
                name,
                phone,
                final_date,
                final_time
            )

            print("\nSusan: Perfect!")

            print(
                "Susan: Your appointment has been booked successfully."
            )

            print(
                f"Susan: Appointment ID: {appointment_id}"
            )

            print(
                f"Susan: Scheduled for {final_date} at {final_time}."
            )

            print(
                "Susan: The appointment has been saved."
            )

        else:

            print(
                "\nSusan: No problem. The appointment was not saved."
            )

        print()

        continue

    # =================================
    # NORMAL CHAT
    # =================================

    prompt = f"""
You are Susan, the receptionist of Gyrobot.

Rules:
- Your name is Susan.
- You work for Gyrobot.
- Be friendly and professional.
- Keep answers concise.
- If the customer wants an appointment,
  tell them to say:
  'I want to book an appointment'

Customer:
{user_message}
"""

    try:

        response = model.generate_content(prompt)

        print("\nSusan:", response.text)

        print()

    except Exception as e:

        print("\nError:", str(e))

